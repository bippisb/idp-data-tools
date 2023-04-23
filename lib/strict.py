from openpyxl import Workbook


def codebook_openpyxl(codebook: Workbook):
    sheet = codebook.get_sheet_by_name('codebook')
    if sheet['A1'] == 'Dataset variables & Formulas used' and sheet['A2'] == 'Variable' and sheet['B2'] == 'Variable description' and sheet['C2'] == 'Variable type' and sheet['D2'] == 'Unit of measurment' and sheet['E2'] == 'Constant Unit/Changing Unit' and sheet['F2'] == 'Formula' and sheet['G2'] == 'Unit Reference' and sheet['H2'] == 'Parent variable' and sheet['I2'] == 'Unit conversion' and sheet['J2'] == 'Original/Derived' and sheet['K2'] == 'Variable Parent':
        print('columns are correct')
    else:
        sheet['A1'] = 'Dataset variables & Formulas used'
        sheet['A2'] = 'Variable name'
        sheet['B2'] = 'Variable description'
        sheet['C2'] = 'Variable type'
        sheet['D2'] = 'Unit of measurment'
        sheet['E2'] = 'Constant Unit/Changing Unit'
        sheet['F2'] = 'Formula'
        sheet['G2'] = 'Unit Reference'
        sheet['H2'] = 'Parent variable'
        sheet['I2'] = 'Unit conversion'
        sheet['J2'] = 'Original/Derived'
        sheet['K2'] = 'Variable Parent'
        print('replaced coloumns')
    return codebook


def metadata_openpyxl(metadata: Workbook):
    sheet = metadata.get_sheet_by_name('metadata information')
    if sheet['A1'] == 'Metadata information' and sheet['A2'] == 'Domain' and sheet['A3'] == 'Dataset name' and sheet['A4'] == 'Granularity level' and sheet['A5'] == 'Frequency' and sheet['A6'] == 'Source name' and sheet['A7'] == 'Source link' and sheet['A8'] == 'Data retrieval date' and sheet['A9'] == 'Data last updated' and sheet['A10'] == 'Data extraction page' and sheet['A11'] == 'About' and sheet['A12'] == 'Methodology' and sheet['A13'] == 'Resource':
        print('columns are correct')
    else:
        sheet['A1'] = 'Metadata information'
        sheet['A2'] = 'Domain'
        sheet['A3'] = 'Dataset name'
        sheet['A4'] = 'Granularity level'
        sheet['A5'] = 'Frequency'
        sheet['A6'] = 'Source name'
        sheet['A7'] = 'Source link'
        sheet['A8'] = 'Data retrieval date'
        sheet['A9'] = 'Data last updated'
        sheet['A10'] = 'Data extraction page'
        sheet['A11'] = 'About'
        sheet['A12'] = 'Methodology'
        sheet['A13'] = 'Resource'
        print('replaced coloumns')
    return metadata


def additional_info_openpyxl(additional_info: Workbook):
    sheet = additional_info.get_sheet_by_name('additional information')
    if sheet['A1'] == 'Additional information' and sheet['A2'] == 'Years covered' and sheet['A3'] == 'Number of State(s) / Union Territories' and sheet['A4'] == 'Number of District(s)' and sheet['A5'] == 'Additional information':
        print('columns are correct')
    else:
        sheet['A1'] = 'Additional information'
        sheet['A2'] = 'Years covered'
        sheet['A3'] = 'Number of State(s) / Union Territories'
        sheet['A4'] = 'Number of District(s)'
        sheet['A5'] = 'Additional information'
        print('replaced coloumns')
    return additional_info


def codebook_value_qa(codebook: Workbook):
    sheet = codebook.get_sheet_by_name('codebook')
    cells = []
    for i in range(3, sheet.max_row + 1):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cell_c = 'C'+str(i)
        cell_d = 'D'+str(i)
        cell_e = 'E'+str(i)
        cell_f = 'F'+str(i)
        cell_g = 'G'+str(i)
        cell_h = 'H'+str(i)
        cell_i = 'I'+str(i)
        cell_j = 'J'+str(i)
        cells = [cell_a, cell_b, cell_c, cell_d, cell_e,
                 cell_f, cell_g, cell_h, cell_i, cell_j]
        for cell in cells:
            if sheet[cell].value == None:
                print(sheet[cell] + 'is empty in codebook sheet')
                pass
            elif type(sheet[cell].value) == int:
                pass
            else:
                sheet[cell].value.title().strip()
                if 'Year' or 'Date' in sheet[cell].value:
                    date = 'C'+str(cell)
                    sheet[date].value = 'Date'
                elif 'Num' in sheet[cell].value:
                    num = 'C'+str(cell)
                    sheet[num].value = 'Numeric'
                    numeric = 'E'+str(cell)
                    sheet[numeric].value = 0
                elif 'state' or 'district' in sheet[cell].value:
                    text = 'C'+str(cell)
                    sheet[text].value = 'Region'
                else:
                    text = 'C'+str(cell)
                    sheet[text].value = 'Categorical'
                    text = 'E'+str(cell)
                    sheet[text].value = ''


def metadata_value_qa(metadata: Workbook):
    sheet = metadata.get_sheet_by_name('metadata information')
    cells = []
    for i in range(3, sheet.max_row + 1):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cells = [cell_a, cell_b]
        for cell in cells:
            if sheet[cell].value == None:
                print(sheet[cell] + 'is empty in metadata sheet')
                pass
            else:
                sheet[cell].value.title().strip()


def additional_info_value_qa(additional_info: Workbook):
    sheet = additional_info.get_sheet_by_name('additional information')
    cells = []
    for i in range(sheet.max_row + 1):
        cell_a = 'A'+str(i)
        cell_b = 'B'+str(i)
        cells = [cell_a, cell_b]
        for cell in cells:
            if sheet[cell].value == None:
                print(sheet[cell] + 'is empty in additional information sheet')
                pass
            else:
                sheet[cell].value.title().strip(). replace('/', '').replace(':', '').replace('*', '').replace(
                    '?', '').replace('"', '').replace('<', '').replace('>', '').replace('|', '')
